import sys

import cv2
from PyQt5.QtCore import pyqtSlot, Qt, QThread, pyqtSignal
from PyQt5.uic import loadUi
from PyQt5 import QtGui
from PyQt5.QtGui import QPixmap
from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import QDialog, QApplication, QWidget, QVBoxLayout, QPushButton, QMainWindow, QMessageBox, QLabel, QFileDialog
import MySQLdb as mdb
from Manager_User import AddUser,UpdateUser
from Manager_History import DetailsHistory
from detect import *;
from cvzone import SerialModule
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem

arduino = SerialModule.SerialObject("COM3")
class VideoThread(QThread):
    change_pixmap_signal = pyqtSignal(np.ndarray)
    change_text_signal = pyqtSignal(str)
    change_pixmap_signal_license = pyqtSignal(np.ndarray)
    change_pixmap_signal_fullpicture = pyqtSignal(np.ndarray)
    def __init__(self):
        super().__init__()
        self._run_flag = True
    def run(self):
        # Tạo một đối tượng VideoCapture để đọc video
        cap = cv2.VideoCapture(0)
    
        # Lấy kích thước khung hình của video
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

        # Đọc từng khung hình của video, xử lý và ghi lại
        result_s = ''
        result_before = ''
        r = 0

        while self._run_flag:
            ret, frame = cap.read()
            if not ret:
                break
            processed_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = model(processed_frame)
            max_value_before = 0
            id = 1
                
            if len(results.pred[0]) >= 1:
                if (results.pred[0][0][3] >= height*7/8 and r == 1):
                    arduino.sendData([0])
                    r = 0
                    self.change_text_signal.emit('')
                for i in range (len(results.pred[0])):
                    accuracy = float(results.pred[0][i][4].item())
                    # print(results.pred)
                    #(results.pred[0][i][2].item() - results.pred[0][i][0].item()) >= 80
                    count = 0
                    if (accuracy >= 0.7 and (results.pred[0][i][2].item() - results.pred[0][i][0].item()) >= 0 and r == 0 and results.pred[0][i][3] < height*7/8):
                        if (readPlate(processed_frame,model) != None):
                            k, result_s, max_value, cropped_image = readPlate(processed_frame,model)
                            print(max_value, " ", result_s)
                            if (max_value < 220):
                                if (max_value >= max_value_before):
                                    if (result_s == result_before):
                                        count = count + 1
                                    result_before = result_s
                                    result_s = ""
                                    if (count == 5): result_s = result_before

                                    max_value_before = max_value
                                else:
                                    result_s = result_before
                                    print(max_value_before)
                        if (result_s != ""):
                            r = 1
                            self.change_pixmap_signal_license.emit(cv2.cvtColor(cropped_image, cv2.COLOR_RGB2BGR))
                            self.change_pixmap_signal_fullpicture.emit(cv2.cvtColor(results.render()[0], cv2.COLOR_RGB2BGR))
                            self.change_text_signal.emit(result_s)
                            # arduino.sendData([1])
                                    # cv2.imwrite(f"{id}.jpg", cropped_image)
                            with open('license_plates_history.txt', 'a') as file:
                                file.write('{} {}\n'.format(id, result_s))
                            id = id + 1
            else:
                r = 0
                self.change_text_signal.emit('')
                arduino.sendData([0])
                
            cv2.putText(results.render()[0], result_s, (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)
            if (r == 0):
                cv2.putText(results.render()[0], "Next license plate!", (50, 200), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)        
                        # cv2.imshow('Processed Frame', )
            self.change_pixmap_signal.emit(cv2.cvtColor(results.render()[0], cv2.COLOR_RGB2BGR))
            # shut down capture system
        cap.release()    
    
    def stop(self):
        """Sets run flag to False and waits for thread to finish"""
        self._run_flag = False
        self.wait()
        
class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('E:/BKDN/Season_3_Term_2/PBL5/pbl5_app/PBL5-Car-Lisence-Plate-Recognition-Application-v2/My-App-PBL5/UI/main.ui', self)
        
        # control
        self.lbCamera
        self.lbCamera.resize(640, 480)
        self.btnOpen.clicked.connect(self.OpenGate)
        self.btnCancel.clicked.connect(self.CloseGate)
        # user
        self.btnAddUser.clicked.connect(self.Add)
        self.btnUpdateUser.clicked.connect(self.Update)
        self.btnDelUser.clicked.connect(self.Delete)
        self.btnSearchUser.clicked.connect(self.SearchUser)
        self.btnResetUser.clicked.connect(self.ResetUser)
        self.btnSortUserIncrease.clicked.connect(self.SortIncreaseUser)
        self.btnSortUserDecrease.clicked.connect(self.SortDecreaseUser)
        self.tableUser.setColumnWidth(0, 100)
        self.tableUser.setColumnWidth(1, 250)
        self.tableUser.setColumnWidth(2, 100)
        self.tableUser.setColumnWidth(3, 200)
        self.tableUser.setColumnWidth(4, 200)
        self.tableUser.setColumnWidth(5, 100)
        self.tableUser.setColumnWidth(6, 250)

        # history
        self.btnDetailsHistory.clicked.connect(self.Details)
        self.btnSearchHistory.clicked.connect(self.SearchHistory)
        self.btnResetHistory.clicked.connect(self.ResetHistory)
        self.btnSortHistoryIncrease.clicked.connect(self.SortIncreaseHistory)
        self.btnSortHistoryDecrease.clicked.connect(self.SortDecreaseHistory)
        self.btnExportExcel.clicked.connect(self.ExportFileExcel)
        self.tableHistory.setColumnWidth(0, 100)
        self.tableHistory.setColumnWidth(1, 100)
        self.tableHistory.setColumnWidth(2, 250)
        self.tableHistory.setColumnWidth(3, 200)
        self.tableHistory.setColumnWidth(4, 200)
        self.tableHistory.setColumnWidth(5, 290)
        # show
        self.ShowListUser()
        self.ShowListHistory()
        #main
        self.btn_CreateNewUser.clicked.connect(self.CreateNewUser)
        # create the video capture thread
        self.thread = VideoThread()
        # connect its signal to the update_image slot
        self.thread.change_pixmap_signal.connect(self.update_image)
        # connect its signal to the update_text slot
        self.thread.change_text_signal.connect(self.update_text)
        # connect its signal to the update_text slot
        self.thread.change_pixmap_signal_license.connect(self.update_license_picture)
        # connect its signal to the update_text slot
        self.thread.change_pixmap_signal_fullpicture.connect(self.update_license_full_picture)
        #
        #self.plate_number = ""
        #self.LoadControl()
        # start the thread
        self.thread.start()
        
    def closeEvent(self, event):
        self.thread.stop()
        event.accept()
    
    def update_text(self, text):
        self.txtLicensePlate.setText(text)
        # self.plate_number = text
        if text != '':
            try:
                db = mdb.connect('localhost', 'root', '', 'pbl5_db')
                cursor = db.cursor()
                query = "SELECT * FROM users WHERE Plate = %s"
                values = (str(text),)
                cursor.execute(query, values)
                row = cursor.fetchone()
                self.txtUsername.setText(row[1])
                self.txtLicensePlate.setText(row[6])
                self.ShowControl(text, self.Save_Temp_license)
            except Exception as e:
                print(f"Lỗi: {str(e)}")
            print('load')
        else: 
            self.txtUsername.setText("")
            self.label_lisence.clear()
    @pyqtSlot(np.ndarray)
    def update_image(self, frame):
        """Updates the image_label with a new opencv image"""
        qt_img = self.convert_cv_qt(frame)
        self.lbCamera.setPixmap(qt_img)
        
    @pyqtSlot(np.ndarray)
    def update_license_picture(self, frame):
        """Updates the image_label with a new opencv image"""
        qt_img = self.convert_cv_qt(frame)
        self.label_lisence.setPixmap(qt_img)

    @pyqtSlot(np.ndarray)
    def update_license_full_picture(self, frame):
        """Updates the image_label with a new opencv image"""
        self.Save_Temp_license = frame
        
    def convert_cv_qt(self, frame):
        """Convert from an opencv image to QPixmap"""
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(640, 480, Qt.KeepAspectRatio)
        return QPixmap.fromImage(p)
    #Control
    def OpenGate(self):
        arduino.sendData([1])
        print('true')
    def CloseGate(self):
        arduino.sendData([0])
    # USER
    #def LoadControl(self, plate_number):
        
    def ShowControl(self , id, frame):
        # plate_number = 5
        plate_number = id
        # Chuyển đổi frame thành một đối tượng bytes
        buffer = BytesIO()
        np.save(buffer, frame)
        frame_bytes = buffer.getvalue()
        
        db = mdb.connect('localhost', 'root', '', 'pbl5_db')
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM users WHERE Plate = %s", (plate_number,))
        plate_exists = cursor.fetchone()[0]
        if plate_exists:
                # lấy id_user
            cursor = db.cursor()
                # Thực thi câu truy vấn SQL để lấy id_user
            cursor.execute("SELECT IdUser FROM users WHERE Plate = %s", (plate_number,))
            id_user = cursor.fetchone()[0]
                # print(id_user)
            try:
                query = "INSERT INTO history (Time,IdUser,HistoryImage) SELECT  CURRENT_TIMESTAMP , %s,%s from users where Plate = %s "
                cursor.execute(query,(id_user, frame_bytes, plate_number,))
                db.commit()
                arduino.sendData([1])
                print('Frame saved to database!')
            except Exception as e:
                print(f"Lỗi Showcontrol: {str(e)}")
        db.close()
    #
    def CreateNewUser(self):
        msgBox = QMessageBox()
        msgBox.setText("Have you added a new user?")
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDefaultButton(QMessageBox.No)
        result = msgBox.exec_()
        if result == QMessageBox.Yes:
            self.Add()
            print('Thêm user')

    #old
    def ShowListUser(self):
        db = mdb.connect('localhost', 'root', '', 'pbl5_db')
        cursor = db.cursor()
        query = "SELECT *FROM users"
        cursor.execute(query)
        rows = cursor.fetchall()
        self.tableUser.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                self.tableUser.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                self.tableUser.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        cursor.close()
        db.close()
    idAction = None
    def Add(self):
        dialog = AddUser()
        if dialog.exec_():
            self.ShowListUser()
    def Update(self):
           selectedRow = self.tableUser.currentRow()
           if selectedRow >= 0:
               id = self.tableUser.item(selectedRow, 0).text()
               int_id =  int(id)
               self.idAction = int_id
               dialog = UpdateUser(self.idAction)
               if dialog.exec_():
                   self.ShowListUser()
    def Delete(self):
        # try:
        selectedRow = self.tableUser.currentRow()
        if selectedRow >= 0:
            id = self.tableUser.item(selectedRow, 0).text()
            # Hiển thị form
            msgBox = QMessageBox()
            msgBox.setText("If you delete this user, its history will be deleted according to ?")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDefaultButton(QMessageBox.No)
            result = msgBox.exec_()
            if result == QMessageBox.Yes:
                id_int = int(id)
                values = (id_int,)
                # Xóa khách hàng khỏi database
                db = mdb.connect('localhost', 'root', '', 'pbl5_db')
                cursor = db.cursor()
                query2 = "DELETE FROM history WHERE IdUser =%s"
                cursor.execute(query2, values)
                db.commit()
                cursor.close()
                db.close()
                db = mdb.connect('localhost', 'root', '', 'pbl5_db')
                cursor = db.cursor()
                self.ShowListHistory()
                query1 = "DELETE FROM users WHERE IdUser = %s"
                cursor.execute(query1, values)
                db.commit()
                cursor.close()
                db.close()
                self.ShowListUser()
    def SearchUser(self):
        try:
            input = self.txtSearchUser.text()
            db = mdb.connect('localhost', 'root', '', 'pbl5_db')
            cursor = db.cursor()
            query = "SELECT * FROM Users WHERE Name like '%" +input +"%' or PhoneNumber like '%"+input+"%' or Plate like'%" \
                    +input +"%' or IdentityCard like'%" + input + "%'"
            cursor.execute(query)
            rows = cursor.fetchall()
            self.tableUser.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.tableUser.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                    self.tableUser.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            cursor.close()
            db.close()

        except Exception as e:
            print(f"Lỗi: {str(e)}")

    def ResetUser(self):
        self.txtSearchUser.clear()
        self.ShowListUser()
    def SortIncreaseUser(self):
        try:
            cbbSelectedValue = self.cbbSortUser.currentText()
            db = mdb.connect('localhost', 'root', '', 'pbl5_db')
            cursor = db.cursor()
            match cbbSelectedValue:
                case "ID":
                    query = "SELECT * FROM Users ORDER BY IdUser ASC"
                case "Name":
                    query = "SELECT * FROM Users ORDER BY Name ASC"
                case "Phone Number":
                    query = "SELECT * FROM Users ORDER BY PhoneNumber ASC"
                case "Lisence Plate":
                    query = "SELECT * FROM Users ORDER BY Plate ASC"
            cursor.execute(query)
            rows = cursor.fetchall()
            self.tableUser.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.tableUser.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                    self.tableUser.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            cursor.close()
            db.close()
        except Exception as e:
            print(f"Lỗi: {str(e)}")
    def SortDecreaseUser(self):
        try:
            cbbSelectedValue = self.cbbSortUser.currentText()
            db = mdb.connect('localhost', 'root', '', 'pbl5_db')
            cursor = db.cursor()
            match cbbSelectedValue:
                case "ID":
                    query = "SELECT * FROM Users ORDER BY IdUser DESC"
                case "Name":
                    query = "SELECT * FROM Users ORDER BY Name DESC"
                case "Phone Number":
                    query = "SELECT * FROM Users ORDER BY PhoneNumber DESC"
                case "Lisence Plate":
                    query = "SELECT * FROM Users ORDER BY Plate DESC"
            cursor.execute(query)
            rows = cursor.fetchall()
            self.tableUser.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.tableUser.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                    self.tableUser.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            cursor.close()
            db.close()
        except Exception as e:
            print(f"Lỗi: {str(e)}")

    # HISTORY
    def ShowListHistory(self):
        db = mdb.connect('localhost', 'root', '', 'pbl5_db')
        cursor = db.cursor()
        query = "SELECT history.IdHistory, users.IdUser ,users.Name, users.Plate, users.PhoneNumber, history.Time FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY history.IdHistory ASC"
        cursor.execute(query)
        rows = cursor.fetchall()
        self.tableHistory.setRowCount(len(rows))

        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                self.tableHistory.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                self.tableHistory.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        cursor.close()
        db.close()

    idAction2 = None
    def Details(self):
        try:
           selectedRow = self.tableHistory.currentRow()
           if selectedRow >= 0:
               id = self.tableHistory.item(selectedRow, 0).text()
               int_id = int(id)
               self.idAction2 = int_id
               dialog = DetailsHistory(self.idAction2)
               if dialog.exec_():
                   self.ShowListHistory()
        except Exception as e:
            print(f"Lỗi: {str(e)}")
    def SearchHistory(self):
        try:
            input = self.txtSearchHistory.text()
            db = mdb.connect('localhost', 'root', '', 'pbl5_db')
            cursor = db.cursor()
            query = "SELECT history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                    "FROM users JOIN history ON users.IdUser = history.IdUser " \
                    "WHERE users.Plate LIKE '%" +input +"%'or Users.Name like'%" + input +"%'" \
                    "or Users.PhoneNumber like '%" + input +"%' or History.Time like '%" + input+ "%';"
            print(query + "\n")
            cursor.execute(query)
            rows = cursor.fetchall()
            self.tableHistory.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.tableHistory.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                    self.tableHistory.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            cursor.close()
            db.close()
        except Exception as e:
            print(f"Lỗi: {str(e)}")

#================== Load lại dữ liệu =======
    def ResetHistory(self):
        self.txtSearchHistory.clear()
        self.ShowListHistory()
# ================ Sắp xếp tăng dần theo Id, tên, số điện thoại, biển số xe, thời gian ra vào ==============
    def SortIncreaseHistory(self):
        try:
            cbbSelectedValue = self.cbbSortHistory.currentText()
            db = mdb.connect('localhost', 'root', '', 'pbl5_db')
            cursor = db.cursor()
            match cbbSelectedValue:
                case "ID History":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY history.IdHistory ASC"
                case "Name":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY Users.name ASC"
                case "Phone Number":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY Users.PhoneNumber ASC"
                case "Lisence Plate":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY Users.Plate ASC"
                case "Time Check-in":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY History.Time ASC"
            cursor.execute(query)
            rows = cursor.fetchall()
            self.tableHistory.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.tableHistory.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                    self.tableHistory.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            cursor.close()
            db.close()
        except Exception as e:
            print(f"Lỗi: {str(e)}")

    # =============== Sắp xếp giảm dần ==================================
    def SortDecreaseHistory(self):
        try:
            cbbSelectedValue = self.cbbSortHistory.currentText()
            db = mdb.connect('localhost', 'root', '', 'pbl5_db')
            cursor = db.cursor()
            match cbbSelectedValue:
                case "ID History":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY history.IdHistory DESC"
                case "Name":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY Users.name DESC"
                case "Phone Number":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY Users.PhoneNumber DESC"
                case "Lisence Plate":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY Users.Plate DESC"
                case "Time Check-in":
                    query = "SELECT  history.IdHistory, users.IdUser, users.Name, users.Plate, users.PhoneNumber, history.Time " \
                            "FROM Users INNER JOIN History ON users.IdUser = history.IdUser ORDER BY History.Time DESC"
            cursor.execute(query)
            rows = cursor.fetchall()
            self.tableHistory.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.tableHistory.setItem(i, j, QtWidgets.QTableWidgetItem(str(val)))
                    self.tableHistory.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            cursor.close()
            db.close()
        except Exception as e:
            print(f"Lỗi: {str(e)}")
    def ExportFileExcel(self):
        # Tạo một đối tượng Workbook
        workbook = Workbook()
        # Tạo một trang tính mới
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = "History ID"
        sheet.cell(row=1, column=2).value = "User ID"
        sheet.cell(row=1, column=3).value = "Name"
        sheet.cell(row=1, column=4).value = "License Plate"
        sheet.cell(row=1, column=5).value = "Phone Number"
        sheet.cell(row=1, column=6).value = "Time Check-in"
        # Duyệt qua từng dòng và cột của QTableWidget để lấy dữ liệu
        for row in range(self.tableHistory.rowCount()):
            for col in range(self.tableHistory.columnCount()):
                item = self.tableHistory.item(row, col)
                if item is not None:
                    # Lấy giá trị từ Qself.tableHistoryItem và gán vào ô tương ứng trong trang tính
                    sheet.cell(row=row+2, column=col+1).value = item.text()
        # Hiển thị hộp thoại lưu tệp tin và lưu Workbook vào đường dẫn được chọn
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_path, _ = QFileDialog.getSaveFileName(None, "Lưu tệp tin Excel", "", "Excel Files (*.xlsx)", options=options)

        if file_path:
            workbook.save(file_path)
